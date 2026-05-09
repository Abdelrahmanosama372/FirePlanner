from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from fireplanner.boq.models import BOQReport
from fireplanner.boq.output.formatting import (
    connection_headers,
    connection_sections,
    paint_headers,
    paint_rows,
    pipe_headers,
    pipe_rows,
)


class BOQExcelExporter:
    @staticmethod
    def export(report: BOQReport, output_path: str | Path) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "BOQ"

        row = 1
        row = BOQExcelExporter._write_section_title(sheet, row, "PIPES")
        row = BOQExcelExporter._write_table(sheet, row, pipe_headers(), pipe_rows(report))
        row = BOQExcelExporter._write_unit(sheet, row, report.pipes.unit.value)
        row += 1

        row = BOQExcelExporter._write_section_title(sheet, row, "CONNECTIONS")
        for section_name, rows in connection_sections(report):
            row = BOQExcelExporter._write_subtitle(sheet, row, section_name)
            row = BOQExcelExporter._write_table(sheet, row, connection_headers(), rows)
            row += 1
        row = BOQExcelExporter._write_unit(sheet, row, report.connections.unit.value)
        row += 1

        row = BOQExcelExporter._write_section_title(sheet, row, "PAINT")
        BOQExcelExporter._write_table(sheet, row, paint_headers(), paint_rows(report))

        BOQExcelExporter._auto_width(sheet)
        path = Path(output_path)
        workbook.save(path)
        return path

    @staticmethod
    def _write_section_title(sheet: Worksheet, row: int, title: str) -> int:
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=14)
        return row + 1

    @staticmethod
    def _write_subtitle(sheet: Worksheet, row: int, title: str) -> int:
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=12)
        return row + 1

    @staticmethod
    def _write_table(
        sheet: Worksheet,
        row: int,
        headers: tuple[str, ...],
        rows: list[tuple[str, ...]],
    ) -> int:
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for data_row in rows:
            for col, value in enumerate(data_row, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="left")
            row += 1

        return row

    @staticmethod
    def _write_unit(sheet: Worksheet, row: int, unit: str) -> int:
        sheet.cell(row=row, column=1, value=f"Unit: {unit}").font = Font(italic=True)
        return row + 1

    @staticmethod
    def _auto_width(sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            max_len = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[column].width = max(12, min(max_len + 2, 60))
